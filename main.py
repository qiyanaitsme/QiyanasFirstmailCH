import openpyxl
import subprocess
import socket
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed

def ping(host):
    try:
        subprocess.check_output(["ping", "-c", "1", "-W", "2", host], stderr=subprocess.STDOUT, universal_newlines=True)
        return True
    except subprocess.CalledProcessError:
        return False

def get_ip(domain):
    try:
        return socket.gethostbyname(domain)
    except socket.gaierror:
        return None

def is_firstmail(ip):
    return ip == '5.252.35.241'

def process_domain(row, domain):
    result = {"row": row, "domain": domain, "ip": None, "is_firstmail": False}
    ip = get_ip(domain)
    if ip:
        result["ip"] = ip
        result["is_firstmail"] = is_firstmail(ip)
    return result

wb = openpyxl.load_workbook('FIRST.xlsx')
sheet = wb.active

green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

if sheet['A1'].value != 'DOMAIN':
    sheet['A1'] = 'DOMAIN'
    sheet['B1'] = 'IP'
    sheet['C1'] = 'FIRSTMAIL'
    sheet['D1'] = 'NOT FIRSTMAIL'

domains = [(row, sheet[f'A{row}'].value) for row in range(2, sheet.max_row + 1) if sheet[f'A{row}'].value]

with ThreadPoolExecutor(max_workers=20) as executor:
    future_to_domain = {executor.submit(process_domain, row, domain): (row, domain) for row, domain in domains}
    for future in as_completed(future_to_domain):
        row, domain = future_to_domain[future]
        try:
            result = future.result()
            if result["ip"]:
                sheet[f'B{row}'] = result["ip"]
                if result["is_firstmail"]:
                    sheet[f'C{row}'].fill = green_fill
                    sheet[f'D{row}'].fill = black_fill
                else:
                    sheet[f'C{row}'].fill = black_fill
                    sheet[f'D{row}'].fill = green_fill
            else:
                sheet[f'B{row}'] = "Не удалось получить IP"
        except Exception:
            sheet[f'B{row}'] = "Ошибка при обработке"

wb.save('FIRST_updated.xlsx')